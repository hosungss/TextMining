from __future__ import annotations

import argparse
import hashlib
import html
import json
import math
import os
import random
import re
import sys
import time
from collections import Counter
from datetime import datetime, timedelta, timezone
from itertools import combinations
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from advanced_statistics import run_advanced_statistics
except ImportError:
    run_advanced_statistics = None


API_BASE = "https://www.googleapis.com/youtube/v3"
APP_VERSION = "v1.0"
AUTHOR_CREDIT = (
    "서강대학교 가상융합전문대학원 메타버스비즈니스 전공 박사과정 손호성 작성"
)
TOKEN_PATTERN = re.compile(r"[가-힣]{2,}|[A-Za-z][A-Za-z0-9+#.-]{1,}")
KOREAN_PATTERN = re.compile(r"[가-힣]")
URL_PATTERN = re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE)
WHITESPACE_PATTERN = re.compile(r"\s+")
FORMULA_PREFIXES = ("=", "+", "-", "@")


class YouTubeAPIError(RuntimeError):
    def __init__(self, resource: str, status: int | None, reason: str, message: str):
        super().__init__(f"YouTube API 오류 [{resource}] {status or ''} {reason}: {message}")
        self.resource = resource
        self.status = status
        self.reason = reason
        self.message = message


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_config(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    with path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)

    required = ["queries", "keyword_groups", "positive_words", "negative_words"]
    missing = [name for name in required if name not in config]
    if missing:
        raise ValueError(f"config.json에 필수 항목이 없습니다: {', '.join(missing)}")
    if not config["queries"]:
        raise ValueError("queries에는 검색어가 하나 이상 필요합니다.")
    if config.get("order", "date") not in {
        "date",
        "rating",
        "relevance",
        "title",
        "viewCount",
    }:
        raise ValueError("order 값은 date, rating, relevance, title, viewCount 중 하나여야 합니다.")
    return config


def save_csv(frame: pd.DataFrame, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def chunks(values: list[str], size: int) -> Iterable[list[str]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def parse_iso8601_duration(value: str | None) -> int:
    if not value:
        return 0
    match = re.fullmatch(
        r"P(?:(?P<days>\d+)D)?(?:T(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?)?",
        value,
    )
    if not match:
        return 0
    parts = {name: safe_int(number) for name, number in match.groupdict().items()}
    return (
        parts["days"] * 86400
        + parts["hours"] * 3600
        + parts["minutes"] * 60
        + parts["seconds"]
    )


def clean_text(value: Any) -> str:
    text = "" if value is None or (isinstance(value, float) and math.isnan(value)) else str(value)
    text = URL_PATTERN.sub(" ", text)
    return WHITESPACE_PATTERN.sub(" ", text).strip()


def tokenise(text: Any, stopwords: set[str], min_length: int = 2) -> list[str]:
    cleaned = clean_text(text).lower()
    tokens = TOKEN_PATTERN.findall(cleaned)
    return [
        token
        for token in tokens
        if len(token) >= min_length and token not in stopwords and not token.isdigit()
    ]


def count_term(text: str, term: str) -> int:
    text_lower = text.lower()
    term_lower = term.lower().strip()
    if not term_lower:
        return 0
    if re.fullmatch(r"[a-z0-9+#. -]+", term_lower):
        pattern = rf"(?<![a-z0-9]){re.escape(term_lower)}(?![a-z0-9])"
        return len(re.findall(pattern, text_lower))
    return text_lower.count(term_lower)


def classify_topic(
    text: Any, keyword_groups: dict[str, list[str]]
) -> tuple[str, str, int]:
    cleaned = clean_text(text)
    scores = {
        group: sum(count_term(cleaned, term) for term in terms)
        for group, terms in keyword_groups.items()
    }
    positive = [(group, score) for group, score in scores.items() if score > 0]
    if not positive:
        return "기타", "기타", 0
    positive.sort(key=lambda item: (-item[1], item[0]))
    primary = positive[0][0]
    labels = "|".join(group for group, _ in positive)
    return primary, labels, positive[0][1]


def classify_sentiment(
    text: Any, positive_words: list[str], negative_words: list[str]
) -> tuple[str, int, int, int]:
    cleaned = clean_text(text)
    positive = sum(count_term(cleaned, word) for word in positive_words)
    negative = sum(count_term(cleaned, word) for word in negative_words)
    score = positive - negative
    if score > 0:
        label = "긍정"
    elif score < 0:
        label = "부정"
    else:
        label = "중립"
    return label, score, positive, negative


def api_get(
    resource: str,
    params: dict[str, Any],
    api_key: str,
    timeout: int = 30,
    retries: int = 3,
) -> dict[str, Any]:
    filtered = {key: value for key, value in params.items() if value not in (None, "")}
    filtered["key"] = api_key
    url = f"{API_BASE}/{resource}?{urlencode(filtered)}"
    request = Request(url, headers={"User-Agent": "startup-youtube-research/1.0"})

    for attempt in range(retries + 1):
        try:
            with urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            reason = "httpError"
            message = raw[:500]
            try:
                payload = json.loads(raw).get("error", {})
                message = payload.get("message", message)
                errors = payload.get("errors") or []
                if errors:
                    reason = errors[0].get("reason", reason)
            except json.JSONDecodeError:
                pass
            if exc.code in {429, 500, 502, 503, 504} and attempt < retries:
                time.sleep(2**attempt)
                continue
            raise YouTubeAPIError(resource, exc.code, reason, message) from exc
        except URLError as exc:
            if attempt < retries:
                time.sleep(2**attempt)
                continue
            raise YouTubeAPIError(resource, None, "networkError", str(exc.reason)) from exc
    raise AssertionError("도달할 수 없는 코드")


def fetch_search_results(
    query: str, config: dict[str, Any], api_key: str
) -> list[str]:
    maximum = safe_int(config.get("max_videos_per_query"), 200)
    results: list[str] = []
    page_token: str | None = None

    while len(results) < maximum:
        params = {
            "part": "snippet",
            "q": query,
            "type": "video",
            "maxResults": min(50, maximum - len(results)),
            "order": config.get("order", "date"),
            "publishedAfter": config.get("published_after"),
            "publishedBefore": config.get("published_before"),
            "regionCode": config.get("region_code"),
            "relevanceLanguage": config.get("relevance_language"),
            "pageToken": page_token,
        }
        payload = api_get("search", params, api_key)
        for item in payload.get("items", []):
            video_id = item.get("id", {}).get("videoId")
            if video_id:
                results.append(video_id)
        page_token = payload.get("nextPageToken")
        if not page_token or not payload.get("items"):
            break
        time.sleep(float(config.get("request_delay_seconds", 0.05)))
    return results[:maximum]


def fetch_video_details(video_ids: list[str], api_key: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for batch in chunks(video_ids, 50):
        payload = api_get(
            "videos",
            {
                "part": "snippet,statistics,contentDetails,status",
                "id": ",".join(batch),
                "maxResults": 50,
            },
            api_key,
        )
        for item in payload.get("items", []):
            snippet = item.get("snippet", {})
            statistics = item.get("statistics", {})
            details = item.get("contentDetails", {})
            status = item.get("status", {})
            records.append(
                {
                    "video_id": item.get("id", ""),
                    "title": snippet.get("title", ""),
                    "description": snippet.get("description", ""),
                    "published_at": snippet.get("publishedAt"),
                    "channel_id": snippet.get("channelId", ""),
                    "channel_title": snippet.get("channelTitle", ""),
                    "tags": "|".join(snippet.get("tags", [])),
                    "category_id": snippet.get("categoryId", ""),
                    "default_language": snippet.get("defaultLanguage", ""),
                    "default_audio_language": snippet.get("defaultAudioLanguage", ""),
                    "duration_iso": details.get("duration", ""),
                    "duration_seconds": parse_iso8601_duration(details.get("duration")),
                    "caption_available": details.get("caption", "false"),
                    "definition": details.get("definition", ""),
                    "view_count": safe_int(statistics.get("viewCount")),
                    "like_count": safe_int(statistics.get("likeCount")),
                    "comment_count": safe_int(statistics.get("commentCount")),
                    "privacy_status": status.get("privacyStatus", ""),
                    "made_for_kids": status.get("madeForKids"),
                    "video_url": f"https://www.youtube.com/watch?v={item.get('id', '')}",
                }
            )
    return records


def fetch_channel_details(channel_ids: list[str], api_key: str) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for batch in chunks(sorted(set(channel_ids)), 50):
        if not batch:
            continue
        payload = api_get(
            "channels",
            {"part": "snippet,statistics", "id": ",".join(batch), "maxResults": 50},
            api_key,
        )
        for item in payload.get("items", []):
            statistics = item.get("statistics", {})
            records[item.get("id", "")] = {
                "channel_subscriber_count": safe_int(statistics.get("subscriberCount")),
                "channel_hidden_subscriber_count": bool(
                    statistics.get("hiddenSubscriberCount", False)
                ),
                "channel_view_count": safe_int(statistics.get("viewCount")),
                "channel_video_count": safe_int(statistics.get("videoCount")),
            }
    return records


def parse_comment_resource(
    item: dict[str, Any],
    video_id: str,
    parent_id: str = "",
    source: str = "top_level",
) -> dict[str, Any]:
    snippet = item.get("snippet", {})
    text = snippet.get("textOriginal") or snippet.get("textDisplay") or ""
    return {
        "comment_id": item.get("id", ""),
        "video_id": video_id,
        "parent_id": parent_id,
        "is_reply": bool(parent_id),
        "source": source,
        "text": clean_text(text),
        "like_count": safe_int(snippet.get("likeCount")),
        "published_at": snippet.get("publishedAt"),
        "updated_at": snippet.get("updatedAt"),
    }


def fetch_all_replies(parent_id: str, video_id: str, api_key: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    page_token: str | None = None
    while True:
        payload = api_get(
            "comments",
            {
                "part": "snippet",
                "parentId": parent_id,
                "maxResults": 100,
                "textFormat": "plainText",
                "pageToken": page_token,
            },
            api_key,
        )
        records.extend(
            parse_comment_resource(item, video_id, parent_id, "full_reply")
            for item in payload.get("items", [])
        )
        page_token = payload.get("nextPageToken")
        if not page_token:
            break
    return records


def fetch_video_comments(
    video_id: str, config: dict[str, Any], api_key: str
) -> list[dict[str, Any]]:
    maximum = safe_int(config.get("max_comments_per_video"), 300)
    include_replies = bool(config.get("include_replies", True))
    all_replies = bool(config.get("collect_all_replies", False))
    records: list[dict[str, Any]] = []
    page_token: str | None = None

    while len(records) < maximum:
        payload = api_get(
            "commentThreads",
            {
                "part": "snippet,replies" if include_replies and not all_replies else "snippet",
                "videoId": video_id,
                "maxResults": min(100, maximum - len(records)),
                "order": config.get("comment_order", "time"),
                "textFormat": "plainText",
                "pageToken": page_token,
            },
            api_key,
        )
        for thread in payload.get("items", []):
            top = thread.get("snippet", {}).get("topLevelComment", {})
            top_record = parse_comment_resource(top, video_id)
            records.append(top_record)
            if len(records) >= maximum:
                break

            if include_replies:
                if all_replies and safe_int(thread.get("snippet", {}).get("totalReplyCount")) > 0:
                    replies = fetch_all_replies(top_record["comment_id"], video_id, api_key)
                else:
                    replies = [
                        parse_comment_resource(
                            reply,
                            video_id,
                            top_record["comment_id"],
                            "inline_reply",
                        )
                        for reply in thread.get("replies", {}).get("comments", [])
                    ]
                remaining = maximum - len(records)
                records.extend(replies[:remaining])

            if len(records) >= maximum:
                break
        page_token = payload.get("nextPageToken")
        if not page_token or not payload.get("items"):
            break
        time.sleep(float(config.get("request_delay_seconds", 0.05)))
    return records[:maximum]


def collect(output_root: str | Path, config: dict[str, Any], api_key: str) -> dict[str, int]:
    if not api_key:
        raise ValueError(
            "YOUTUBE_API_KEY 환경변수가 없습니다. API 키 없이 시험하려면 demo 명령을 사용하세요."
        )
    root = Path(output_root)
    raw_dir = root / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    collected_at = utc_now_iso()

    video_matches: dict[str, dict[str, Any]] = {}
    for query in config["queries"]:
        print(f"[수집] 검색어: {query}", flush=True)
        for rank, video_id in enumerate(fetch_search_results(query, config, api_key), start=1):
            entry = video_matches.setdefault(video_id, {"queries": [], "ranks": {}})
            if query not in entry["queries"]:
                entry["queries"].append(query)
            entry["ranks"][query] = rank

    video_ids = list(video_matches)
    print(f"[수집] 중복 제거 후 영상 {len(video_ids):,}개", flush=True)
    video_records = fetch_video_details(video_ids, api_key)
    channel_map = fetch_channel_details(
        [record["channel_id"] for record in video_records], api_key
    )
    for record in video_records:
        match = video_matches.get(record["video_id"], {"queries": [], "ranks": {}})
        record["matched_queries"] = "|".join(match["queries"])
        record["search_ranks_json"] = json.dumps(match["ranks"], ensure_ascii=False)
        record.update(channel_map.get(record["channel_id"], {}))
        record["collected_at_utc"] = collected_at

    videos = pd.DataFrame(video_records)
    save_csv(videos, raw_dir / "videos.csv")

    comment_records: list[dict[str, Any]] = []
    error_records: list[dict[str, Any]] = []
    for index, record in enumerate(video_records, start=1):
        video_id = record["video_id"]
        if safe_int(record.get("comment_count")) <= 0:
            continue
        print(f"[댓글] {index}/{len(video_records)} {video_id}", flush=True)
        try:
            comments = fetch_video_comments(video_id, config, api_key)
            for comment in comments:
                comment["collected_at_utc"] = collected_at
            comment_records.extend(comments)
        except YouTubeAPIError as exc:
            error_records.append(
                {
                    "video_id": video_id,
                    "stage": "comments",
                    "status": exc.status,
                    "reason": exc.reason,
                    "message": exc.message,
                }
            )

    comments = pd.DataFrame(
        comment_records,
        columns=[
            "comment_id",
            "video_id",
            "parent_id",
            "is_reply",
            "source",
            "text",
            "like_count",
            "published_at",
            "updated_at",
            "collected_at_utc",
        ],
    )
    errors = pd.DataFrame(
        error_records, columns=["video_id", "stage", "status", "reason", "message"]
    )
    save_csv(comments, raw_dir / "comments.csv")
    save_csv(errors, raw_dir / "collection_errors.csv")
    (raw_dir / "collection_manifest.json").write_text(
        json.dumps(
            {
                "collected_at_utc": collected_at,
                "queries": config["queries"],
                "video_count": len(videos),
                "comment_count": len(comments),
                "error_count": len(errors),
                "config": config,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return {
        "video_count": len(videos),
        "comment_count": len(comments),
        "error_count": len(errors),
    }


def ensure_columns(frame: pd.DataFrame, defaults: dict[str, Any]) -> pd.DataFrame:
    frame = frame.copy()
    for name, default in defaults.items():
        if name not in frame.columns:
            frame[name] = default
    return frame


def safe_divide(numerator: pd.Series, denominator: pd.Series, scale: float = 1.0) -> pd.Series:
    numerator = pd.to_numeric(numerator, errors="coerce").fillna(0.0)
    denominator = pd.to_numeric(denominator, errors="coerce").fillna(0.0)
    return pd.Series(
        np.where(denominator > 0, numerator / denominator * scale, np.nan),
        index=numerator.index,
        dtype=float,
    )


def duration_band(seconds: Any) -> str:
    value = safe_int(seconds)
    if value < 60:
        return "1분 미만"
    if value < 300:
        return "1~5분"
    if value < 900:
        return "5~15분"
    if value < 1800:
        return "15~30분"
    return "30분 이상"


def bootstrap_ci(
    values: Iterable[float],
    statistic: Callable[[np.ndarray], float],
    iterations: int,
    rng: np.random.Generator,
) -> tuple[float, float]:
    array = np.asarray(list(values), dtype=float)
    array = array[np.isfinite(array)]
    if len(array) == 0:
        return math.nan, math.nan
    if len(array) == 1 or iterations <= 1:
        value = float(statistic(array))
        return value, value
    estimates = np.empty(iterations, dtype=float)
    for index in range(iterations):
        sample = rng.choice(array, size=len(array), replace=True)
        estimates[index] = statistic(sample)
    return tuple(np.quantile(estimates, [0.025, 0.975]).astype(float))


def descriptive_summary(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    labels = {
        "view_count": "조회수",
        "like_count": "좋아요 수",
        "comment_count": "댓글 수",
        "duration_minutes": "영상 길이(분)",
        "age_days": "게시 후 경과일",
        "views_per_day": "일평균 조회수",
        "likes_per_1000_views": "조회 1천회당 좋아요",
        "comments_per_1000_views": "조회 1천회당 댓글",
        "engagement_per_1000_views": "조회 1천회당 참여",
    }
    records: list[dict[str, Any]] = []
    for column in columns:
        values = pd.to_numeric(frame[column], errors="coerce").dropna()
        records.append(
            {
                "metric": labels.get(column, column),
                "column": column,
                "n": len(values),
                "mean": values.mean() if len(values) else np.nan,
                "std": values.std(ddof=1) if len(values) > 1 else np.nan,
                "min": values.min() if len(values) else np.nan,
                "p25": values.quantile(0.25) if len(values) else np.nan,
                "median": values.median() if len(values) else np.nan,
                "p75": values.quantile(0.75) if len(values) else np.nan,
                "max": values.max() if len(values) else np.nan,
            }
        )
    return pd.DataFrame(records)


def group_summary(
    frame: pd.DataFrame,
    group_column: str,
    iterations: int,
    rng: np.random.Generator,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    if frame.empty or group_column not in frame:
        return pd.DataFrame()
    for group, subset in frame.groupby(group_column, dropna=False):
        views = pd.to_numeric(subset["views_per_day"], errors="coerce").dropna()
        engagement = pd.to_numeric(
            subset["engagement_per_1000_views"], errors="coerce"
        ).dropna()
        view_low, view_high = bootstrap_ci(views, np.median, iterations, rng)
        engage_low, engage_high = bootstrap_ci(
            engagement, np.median, iterations, rng
        )
        records.append(
            {
                group_column: group,
                "video_count": len(subset),
                "total_views": subset["view_count"].sum(),
                "median_views": subset["view_count"].median(),
                "median_views_per_day": views.median() if len(views) else np.nan,
                "views_per_day_ci_low": view_low,
                "views_per_day_ci_high": view_high,
                "median_engagement_per_1000": engagement.median()
                if len(engagement)
                else np.nan,
                "engagement_ci_low": engage_low,
                "engagement_ci_high": engage_high,
                "median_duration_minutes": subset["duration_minutes"].median(),
                "median_comment_count": subset["comment_count"].median(),
            }
        )
    return pd.DataFrame(records).sort_values("video_count", ascending=False)


def spearman_with_ci(
    frame: pd.DataFrame,
    pairs: list[tuple[str, str]],
    iterations: int,
    rng: np.random.Generator,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for x_column, y_column in pairs:
        pair = frame[[x_column, y_column]].apply(pd.to_numeric, errors="coerce").dropna()
        n = len(pair)
        if n < 3 or pair[x_column].nunique() < 2 or pair[y_column].nunique() < 2:
            correlation = low = high = np.nan
        else:
            correlation = pair[x_column].rank().corr(pair[y_column].rank())
            estimates: list[float] = []
            array = pair.to_numpy(dtype=float)
            for _ in range(iterations):
                indices = rng.integers(0, n, size=n)
                sample = pd.DataFrame(array[indices], columns=[x_column, y_column])
                if sample[x_column].nunique() < 2 or sample[y_column].nunique() < 2:
                    continue
                estimate = sample[x_column].rank().corr(sample[y_column].rank())
                if np.isfinite(estimate):
                    estimates.append(float(estimate))
            if estimates:
                low, high = np.quantile(estimates, [0.025, 0.975])
            else:
                low = high = np.nan
        records.append(
            {
                "x": x_column,
                "y": y_column,
                "n": n,
                "spearman_rho": correlation,
                "ci_low": low,
                "ci_high": high,
                "note": "부트스트랩 95% CI; 인과관계가 아닌 단순 연관성",
            }
        )
    return pd.DataFrame(records)


def trend_estimates(monthly: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    if monthly.empty or len(monthly) < 2:
        return pd.DataFrame(
            columns=["outcome", "months", "slope_per_month", "r_squared", "direction"]
        )
    x = np.arange(len(monthly), dtype=float)
    for column in ["video_count", "median_views_per_day", "median_engagement_per_1000"]:
        values = pd.to_numeric(monthly[column], errors="coerce")
        valid = values.notna().to_numpy()
        if valid.sum() < 2:
            continue
        slope, intercept = np.polyfit(x[valid], values.to_numpy()[valid], 1)
        predicted = slope * x[valid] + intercept
        observed = values.to_numpy()[valid]
        denominator = np.sum((observed - observed.mean()) ** 2)
        r_squared = (
            1 - np.sum((observed - predicted) ** 2) / denominator
            if denominator > 0
            else np.nan
        )
        direction = "증가" if slope > 0 else "감소" if slope < 0 else "변화 없음"
        records.append(
            {
                "outcome": column,
                "months": int(valid.sum()),
                "slope_per_month": slope,
                "r_squared": r_squared,
                "direction": direction,
                "note": "단순 선형 추세; 마지막 달이 불완전하면 해석 주의",
            }
        )
    return pd.DataFrame(records)


def bootstrap_regression(
    videos: pd.DataFrame,
    iterations: int,
    rng: np.random.Generator,
) -> pd.DataFrame:
    candidate_predictors = {
        "log_age_days": np.log1p(pd.to_numeric(videos["age_days"], errors="coerce")),
        "log_channel_subscribers": np.log1p(
            pd.to_numeric(videos["channel_subscriber_count"], errors="coerce").fillna(0)
        ),
        "log_duration_seconds": np.log1p(
            pd.to_numeric(videos["duration_seconds"], errors="coerce").fillna(0)
        ),
        "title_keyword_mentions": pd.to_numeric(
            videos["title_keyword_mentions"], errors="coerce"
        ).fillna(0),
        "english_startup_in_title": videos["title"].fillna("").str.contains(
            r"\bstartups?\b", case=False, regex=True
        ).astype(float),
        "korean_startup_in_title": videos["title"].fillna("").str.contains(
            "스타트업", regex=False
        ).astype(float),
    }
    design = pd.DataFrame(candidate_predictors, index=videos.index)
    outcome = np.log1p(pd.to_numeric(videos["view_count"], errors="coerce"))
    combined = design.assign(log_views=outcome).replace([np.inf, -np.inf], np.nan).dropna()
    if len(combined) < 20:
        return pd.DataFrame(
            [{"predictor": "분석 불가", "note": "회귀분석에는 완전한 영상 20개 이상이 필요합니다."}]
        )

    predictor_names = [
        name for name in design.columns if combined[name].std(ddof=0) > 0
    ]
    x_raw = combined[predictor_names].to_numpy(dtype=float)
    means = x_raw.mean(axis=0)
    stds = x_raw.std(axis=0)
    x_standardized = (x_raw - means) / stds
    x = np.column_stack([np.ones(len(x_standardized)), x_standardized])
    y = combined["log_views"].to_numpy(dtype=float)
    coefficients = np.linalg.lstsq(x, y, rcond=None)[0]
    predicted = x @ coefficients
    denominator = np.sum((y - y.mean()) ** 2)
    r_squared = 1 - np.sum((y - predicted) ** 2) / denominator if denominator > 0 else np.nan

    bootstrap_values: list[np.ndarray] = []
    for _ in range(iterations):
        indices = rng.integers(0, len(x), size=len(x))
        try:
            estimate = np.linalg.lstsq(x[indices], y[indices], rcond=None)[0]
            if np.all(np.isfinite(estimate)):
                bootstrap_values.append(estimate)
        except np.linalg.LinAlgError:
            continue
    if bootstrap_values:
        bootstrap_array = np.vstack(bootstrap_values)
        ci_low = np.quantile(bootstrap_array, 0.025, axis=0)
        ci_high = np.quantile(bootstrap_array, 0.975, axis=0)
    else:
        ci_low = np.repeat(np.nan, len(coefficients))
        ci_high = np.repeat(np.nan, len(coefficients))

    names = ["intercept"] + predictor_names
    return pd.DataFrame(
        {
            "predictor": names,
            "coefficient": coefficients,
            "ci_low": ci_low,
            "ci_high": ci_high,
            "n": len(combined),
            "model_r_squared": r_squared,
            "interpretation": [
                "절편"
                if name == "intercept"
                else "예측변수 1표준편차 증가와 연관된 log(조회수+1)의 변화"
                for name in names
            ],
            "note": "탐색적 OLS와 부트스트랩 95% CI; 인과효과 아님",
        }
    )


def word_frequency(
    texts: Iterable[Any], stopwords: set[str], min_length: int, limit: int = 200
) -> pd.DataFrame:
    term_frequency: Counter[str] = Counter()
    document_frequency: Counter[str] = Counter()
    document_count = 0
    for text in texts:
        tokens = tokenise(text, stopwords, min_length)
        if not tokens:
            continue
        document_count += 1
        term_frequency.update(tokens)
        document_frequency.update(set(tokens))
    records = [
        {
            "word": word,
            "term_frequency": count,
            "document_frequency": document_frequency[word],
            "document_share": document_frequency[word] / document_count
            if document_count
            else np.nan,
        }
        for word, count in term_frequency.most_common(limit)
    ]
    return pd.DataFrame(
        records,
        columns=["word", "term_frequency", "document_frequency", "document_share"],
    )


def cooccurrence_table(
    texts: Iterable[Any], stopwords: set[str], min_length: int, limit: int = 200
) -> pd.DataFrame:
    tokenized = [tokenise(text, stopwords, min_length) for text in texts]
    overall = Counter(token for tokens in tokenized for token in tokens)
    vocabulary = {word for word, _ in overall.most_common(60)}
    pair_counts: Counter[tuple[str, str]] = Counter()
    for tokens in tokenized:
        selected = sorted(set(tokens) & vocabulary)
        pair_counts.update(combinations(selected, 2))
    records = [
        {"word_1": pair[0], "word_2": pair[1], "cooccurrence": count}
        for pair, count in pair_counts.most_common(limit)
        if count >= 2
    ]
    return pd.DataFrame(records, columns=["word_1", "word_2", "cooccurrence"])


def enrich_videos(videos: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    videos = ensure_columns(
        videos,
        {
            "video_id": "",
            "title": "",
            "description": "",
            "published_at": "",
            "matched_queries": "",
            "channel_id": "",
            "channel_title": "",
            "channel_subscriber_count": 0,
            "duration_seconds": 0,
            "view_count": 0,
            "like_count": 0,
            "comment_count": 0,
            "collected_at_utc": utc_now_iso(),
        },
    )
    for column in [
        "view_count",
        "like_count",
        "comment_count",
        "duration_seconds",
        "channel_subscriber_count",
    ]:
        videos[column] = pd.to_numeric(videos[column], errors="coerce").fillna(0)

    videos["published_datetime"] = pd.to_datetime(
        videos["published_at"], errors="coerce", utc=True
    )
    collected = pd.to_datetime(videos["collected_at_utc"], errors="coerce", utc=True)
    fallback = pd.Timestamp.now(tz="UTC")
    collected = collected.fillna(fallback)
    videos["age_days"] = (
        (collected - videos["published_datetime"]).dt.total_seconds() / 86400
    ).clip(lower=1)
    videos["publish_month"] = videos["published_datetime"].dt.strftime("%Y-%m")
    videos["duration_minutes"] = videos["duration_seconds"] / 60
    videos["duration_band"] = videos["duration_seconds"].map(duration_band)
    videos["views_per_day"] = safe_divide(videos["view_count"], videos["age_days"])
    videos["likes_per_1000_views"] = safe_divide(
        videos["like_count"], videos["view_count"], 1000
    )
    videos["comments_per_1000_views"] = safe_divide(
        videos["comment_count"], videos["view_count"], 1000
    )
    videos["engagement_per_1000_views"] = safe_divide(
        videos["like_count"] + videos["comment_count"], videos["view_count"], 1000
    )
    videos["query_group"] = videos["matched_queries"].fillna("").map(
        lambda value: "복수 키워드"
        if len([part for part in str(value).split("|") if part]) > 1
        else (str(value).split("|")[0] if str(value).strip() else "검색어 미상")
    )
    videos["title_has_korean"] = videos["title"].fillna("").map(
        lambda value: bool(KOREAN_PATTERN.search(str(value)))
    )

    topic_results = videos.apply(
        lambda row: classify_topic(
            f"{row.get('title', '')} {row.get('description', '')}",
            config["keyword_groups"],
        ),
        axis=1,
    )
    videos[["topic_primary", "topic_labels", "topic_score"]] = pd.DataFrame(
        topic_results.tolist(), index=videos.index
    )
    videos["title_keyword_mentions"] = videos["title"].fillna("").map(
        lambda text: sum(count_term(str(text), query) for query in config["queries"])
    )
    videos["description_keyword_mentions"] = videos["description"].fillna("").map(
        lambda text: sum(count_term(str(text), query) for query in config["queries"])
    )
    return videos


def enrich_comments(comments: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    comments = ensure_columns(
        comments,
        {
            "comment_id": "",
            "video_id": "",
            "parent_id": "",
            "is_reply": False,
            "source": "top_level",
            "text": "",
            "like_count": 0,
            "published_at": "",
        },
    )
    comments["like_count"] = pd.to_numeric(comments["like_count"], errors="coerce").fillna(0)
    comments["published_datetime"] = pd.to_datetime(
        comments["published_at"], errors="coerce", utc=True
    )
    comments["comment_month"] = comments["published_datetime"].dt.strftime("%Y-%m")
    sentiment = comments["text"].fillna("").map(
        lambda text: classify_sentiment(
            text, config["positive_words"], config["negative_words"]
        )
    )
    comments[
        ["sentiment", "sentiment_score", "positive_hits", "negative_hits"]
    ] = pd.DataFrame(sentiment.tolist(), index=comments.index)
    topics = comments["text"].fillna("").map(
        lambda text: classify_topic(text, config["keyword_groups"])
    )
    comments[["topic_primary", "topic_labels", "topic_score"]] = pd.DataFrame(
        topics.tolist(), index=comments.index
    )
    comments["startup_keyword_mentions"] = comments["text"].fillna("").map(
        lambda text: sum(count_term(str(text), query) for query in config["queries"])
    )
    return comments


def monthly_video_summary(videos: pd.DataFrame) -> pd.DataFrame:
    valid = videos.dropna(subset=["publish_month"]).copy()
    if valid.empty:
        return pd.DataFrame(
            columns=[
                "publish_month",
                "video_count",
                "total_views",
                "median_views",
                "median_views_per_day",
                "median_engagement_per_1000",
                "total_comments",
            ]
        )
    return (
        valid.groupby("publish_month", as_index=False)
        .agg(
            video_count=("video_id", "nunique"),
            total_views=("view_count", "sum"),
            median_views=("view_count", "median"),
            median_views_per_day=("views_per_day", "median"),
            median_engagement_per_1000=("engagement_per_1000_views", "median"),
            total_comments=("comment_count", "sum"),
        )
        .sort_values("publish_month")
    )


def channel_summary(videos: pd.DataFrame) -> pd.DataFrame:
    if videos.empty:
        return pd.DataFrame()
    return (
        videos.groupby(["channel_id", "channel_title"], dropna=False, as_index=False)
        .agg(
            video_count=("video_id", "nunique"),
            total_views=("view_count", "sum"),
            median_views=("view_count", "median"),
            median_views_per_day=("views_per_day", "median"),
            median_engagement_per_1000=("engagement_per_1000_views", "median"),
            channel_subscriber_count=("channel_subscriber_count", "max"),
        )
        .sort_values(["video_count", "total_views"], ascending=False)
    )


def sentiment_summaries(comments: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if comments.empty:
        return (
            pd.DataFrame(columns=["sentiment", "comment_count", "share"]),
            pd.DataFrame(
                columns=[
                    "comment_month",
                    "comment_count",
                    "positive_share",
                    "negative_share",
                    "neutral_share",
                ]
            ),
        )
    overall = comments.groupby("sentiment", as_index=False).agg(
        comment_count=("comment_id", "count")
    )
    overall["share"] = overall["comment_count"] / overall["comment_count"].sum()

    valid = comments.dropna(subset=["comment_month"])
    counts = (
        valid.groupby(["comment_month", "sentiment"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for label in ["긍정", "부정", "중립"]:
        if label not in counts:
            counts[label] = 0
    counts["comment_count"] = counts[["긍정", "부정", "중립"]].sum(axis=1)
    for label, english in [("긍정", "positive"), ("부정", "negative"), ("중립", "neutral")]:
        counts[f"{english}_share"] = safe_divide(counts[label], counts["comment_count"])
    monthly = counts[
        [
            "comment_month",
            "comment_count",
            "positive_share",
            "negative_share",
            "neutral_share",
        ]
    ].sort_values("comment_month")
    return overall, monthly


def data_quality_table(videos: pd.DataFrame, comments: pd.DataFrame) -> pd.DataFrame:
    records = [
        ("영상 행 수", len(videos)),
        ("고유 영상 수", videos["video_id"].nunique()),
        ("중복 영상 ID 수", int(videos["video_id"].duplicated().sum())),
        ("제목 누락", int(videos["title"].fillna("").eq("").sum())),
        ("게시일 누락", int(videos["published_datetime"].isna().sum())),
        ("조회수 0 또는 누락", int((videos["view_count"] <= 0).sum())),
        ("댓글 행 수", len(comments)),
        ("고유 댓글 수", comments["comment_id"].nunique() if not comments.empty else 0),
        (
            "중복 댓글 ID 수",
            int(comments["comment_id"].duplicated().sum()) if not comments.empty else 0,
        ),
        ("댓글 텍스트 누락", int(comments["text"].fillna("").eq("").sum()) if not comments.empty else 0),
        ("답글 수", int(comments["is_reply"].fillna(False).astype(bool).sum()) if not comments.empty else 0),
    ]
    return pd.DataFrame(records, columns=["check", "value"])


def excel_safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[column]):
            result[column] = result[column].astype(str)
        elif result[column].dtype == "object":
            result[column] = result[column].map(
                lambda value: f"'{value}"
                if isinstance(value, str) and value.startswith(FORMULA_PREFIXES)
                else value
            )
    return result


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1 and sheet.max_column > 0:
        sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        sample = list(column_cells[: min(sheet.max_row, 200)])
        width = min(max(len(str(cell.value or "")) for cell in sample) + 2, 45)
        sheet.column_dimensions[letter].width = max(width, 10)


def write_excel_report(path: Path, tables: dict[str, pd.DataFrame], config: dict[str, Any]) -> None:
    sheet_map = {
        "overview": "개요",
        "descriptive": "기술통계",
        "monthly_videos": "월별영상",
        "trend": "추세",
        "query": "검색어비교",
        "topic": "영상주제",
        "duration": "영상길이",
        "channels": "채널",
        "sentiment": "댓글감성",
        "monthly_comments": "월별댓글",
        "correlations": "상관분석",
        "regression": "조회수회귀",
        "word_video": "영상빈출어",
        "word_comment": "댓글빈출어",
        "cooccurrence": "댓글연관어",
        "top_videos": "상위영상",
        "quality": "데이터품질",
        "videos": "영상자료",
        "comments": "댓글자료",
        "normality_tests": "고급_정규성",
        "variance_tests": "고급_등분산",
        "two_group_tests": "고급_두집단",
        "multi_group_tests": "고급_다집단",
        "posthoc_tests": "고급_사후검정",
        "full_correlations": "고급_상관",
        "contingency_tests": "고급_범주형",
        "regression_models": "고급_회귀모형",
        "regression_diagnostics": "고급_회귀진단",
        "time_series_tests": "고급_시계열",
        "monthly_extended": "고급_월별확장",
        "pca_loadings": "고급_PCA적재량",
        "pca_variance": "고급_PCA분산",
        "pca_cluster_scores": "고급_군집점수",
        "cluster_profiles": "고급_군집프로필",
        "outlier_summary": "고급_이상치",
        "power_analysis": "고급_검정력",
        "method_catalog": "고급_방법론목록",
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overview = pd.DataFrame(
            [
                ["분석 검색어", ", ".join(config["queries"])],
                ["작성", AUTHOR_CREDIT],
                ["버전", APP_VERSION],
                ["분석 기간 시작", config.get("published_after")],
                ["분석 기간 종료", config.get("published_before") or "수집 시점"],
                ["지역/언어", f"{config.get('region_code', '')} / {config.get('relevance_language', '')}"],
                ["감성분석", "사용자 사전 기반 탐색적 분류"],
                ["신뢰구간", f"부트스트랩 {config.get('bootstrap_iterations', 500)}회, 95% CI"],
                ["주의", "검색 결과는 무작위 표본이 아니며 회귀계수는 인과효과가 아님"],
            ],
            columns=["항목", "내용"],
        )
        tables_to_write = {"overview": overview, **tables}
        for key, frame in tables_to_write.items():
            if key not in sheet_map:
                continue
            limited = frame.head(1_000_000)
            excel_safe_frame(limited).to_excel(
                writer, sheet_name=sheet_map[key], index=False
            )

        workbook = writer.book
        for sheet in workbook.worksheets:
            style_sheet(sheet)

        if not tables["monthly_videos"].empty:
            sheet = workbook[sheet_map["monthly_videos"]]
            chart = LineChart()
            chart.title = "월별 게시 영상 수"
            chart.y_axis.title = "영상 수"
            chart.x_axis.title = "월"
            data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            sheet.add_chart(chart, "J2")

        if not tables["topic"].empty:
            sheet = workbook[sheet_map["topic"]]
            chart = BarChart()
            chart.title = "영상 주제별 표본 수"
            data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            sheet.add_chart(chart, "M2")

        if not tables["sentiment"].empty:
            sheet = workbook[sheet_map["sentiment"]]
            chart = PieChart()
            chart.title = "댓글 감성 구성"
            data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 12
            sheet.add_chart(chart, "E2")


MARKDOWN_TABLE_TITLES = {
    "descriptive": "기술통계",
    "monthly_videos": "월별 영상 경향",
    "trend": "영상 게시 추세 추정",
    "query": "검색어별 비교",
    "topic": "주제별 비교",
    "duration": "영상 길이별 비교",
    "channels": "채널별 요약",
    "sentiment": "댓글 감성 구성",
    "monthly_comments": "월별 댓글 경향",
    "correlations": "Spearman 상관분석",
    "regression": "조회수 회귀분석",
    "word_video": "영상 텍스트 빈출어",
    "word_comment": "댓글 빈출어",
    "cooccurrence": "댓글 단어 동시출현",
    "top_videos": "일평균 조회수 상위 영상",
    "quality": "데이터 품질 점검",
    "normality_tests": "정규성 검정",
    "variance_tests": "등분산성 검정",
    "two_group_tests": "두 집단 비교",
    "multi_group_tests": "다집단 비교",
    "posthoc_tests": "사후검정",
    "full_correlations": "확장 상관분석",
    "contingency_tests": "범주형 자료 검정",
    "regression_models": "회귀모형 비교",
    "regression_diagnostics": "회귀 진단",
    "time_series_tests": "시계열 검정",
    "monthly_extended": "월별 확장 지표",
    "pca_loadings": "PCA 적재량",
    "pca_variance": "PCA 설명분산",
    "pca_cluster_scores": "군집 수 평가",
    "cluster_profiles": "군집 프로필",
    "outlier_summary": "이상치 진단",
    "power_analysis": "검정력·필요 표본크기",
    "method_catalog": "통계 방법론 목록",
}

MARKDOWN_TABLE_ORDER = list(MARKDOWN_TABLE_TITLES)

MARKDOWN_TABLE_ALIASES = {
    "descriptive_statistics": "descriptive",
    "monthly_video_trends": "monthly_videos",
    "trend_estimates": "trend",
    "query_comparison": "query",
    "topic_summary": "topic",
    "duration_summary": "duration",
    "channel_summary": "channels",
    "sentiment_summary": "sentiment",
    "monthly_comment_trends": "monthly_comments",
    "spearman_correlations": "correlations",
    "view_regression": "regression",
    "video_word_frequency": "word_video",
    "comment_word_frequency": "word_comment",
    "comment_cooccurrence": "cooccurrence",
    "data_quality": "quality",
}


def markdown_cell(value: Any) -> str:
    if value is None:
        return ""
    try:
        if bool(pd.isna(value)):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, (float, np.floating)):
        text = f"{float(value):.8g}"
    else:
        text = str(value)
    return (
        text.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", "<br>")
        .replace("\n", "<br>")
        .replace("\r", "<br>")
    )


def dataframe_to_markdown(frame: pd.DataFrame) -> str:
    if frame.empty:
        return "_현재 표본과 설정에서 생성된 행이 없습니다._"
    columns = [markdown_cell(column) for column in frame.columns]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in frame.itertuples(index=False, name=None):
        lines.append("| " + " | ".join(markdown_cell(value) for value in row) + " |")
    return "\n".join(lines)


def markdown_table_key(key: str) -> str:
    plain_key = key.removeprefix("advanced_")
    return MARKDOWN_TABLE_ALIASES.get(plain_key, plain_key)


def write_markdown_report(
    path: Path,
    tables: dict[str, pd.DataFrame],
    manifest: dict[str, Any],
    config: dict[str, Any] | None = None,
) -> None:
    config = config or {}
    author_credit = manifest.get("author_credit", AUTHOR_CREDIT)
    app_version = manifest.get("app_version", APP_VERSION)
    queries = manifest.get("queries") or config.get("queries", [])
    start_date = manifest.get("published_after") or config.get("published_after") or "미지정"
    end_date = manifest.get("published_before") or config.get("published_before") or "수집 시점"
    overview = pd.DataFrame(
        [
            ["작성", author_credit],
            ["버전", app_version],
            ["생성 시각(UTC)", manifest.get("generated_at_utc", "")],
            ["검색어", ", ".join(str(query) for query in queries)],
            ["분석 기간", f"{start_date} ~ {end_date}"],
            ["영상 수", manifest.get("video_count", 0)],
            ["댓글 수", manifest.get("comment_count", 0)],
            ["채널 수", manifest.get("unique_channel_count", 0)],
            ["부트스트랩 반복", manifest.get("bootstrap_iterations", config.get("bootstrap_iterations", ""))],
            ["가상자료", "예" if manifest.get("demo_data") else "아니오"],
        ],
        columns=["항목", "내용"],
    )

    canonical_tables: dict[str, pd.DataFrame] = {}
    for key, frame in tables.items():
        canonical_key = markdown_table_key(key)
        if canonical_key in {"videos", "comments", "videos_enriched", "comments_enriched"}:
            continue
        canonical_tables[canonical_key] = frame
    ordered_keys = [key for key in MARKDOWN_TABLE_ORDER if key in canonical_tables]
    ordered_keys.extend(sorted(set(canonical_tables) - set(ordered_keys)))

    lines = [
        "# YouTube 영상·댓글 텍스트 마이닝 전체 결과",
        "",
        f"**{author_credit} · {app_version}**",
        "",
        "> 이 문서는 한 번의 분석에서 생성된 모든 분석 결과표를 통합한 Markdown 보고서입니다. ",
        "> 영상·댓글 정제 원자료 전체는 Excel의 `영상자료`·`댓글자료` 시트와 전체 ZIP 파일에 포함됩니다.",
        "",
        "## 1. 분석 개요",
        "",
        dataframe_to_markdown(overview),
        "",
        "## 2. 분석 및 해석 원칙",
        "",
        "- YouTube 검색 결과는 확률표본이 아니므로 전체 이용자나 전체 영상을 대표한다고 단정할 수 없습니다.",
        "- p값은 Holm 보정 p값, FDR q값, 효과크기, 신뢰구간과 함께 해석합니다.",
        "- 감성 결과는 사용자가 설정한 단어 사전 기반의 탐색적 분류입니다.",
        "- 상관계수와 회귀계수는 변수 간 연관성을 나타내며 인과효과를 의미하지 않습니다.",
        "- 공개 댓글 원문이 포함된 Excel·ZIP을 공유할 때에는 개인정보와 연구윤리를 확인해야 합니다.",
        "",
        "## 3. 결과표 목록",
        "",
        dataframe_to_markdown(
            pd.DataFrame(
                [
                    [
                        index,
                        MARKDOWN_TABLE_TITLES.get(key, key.replace("_", " ")),
                        len(canonical_tables[key]),
                    ]
                    for index, key in enumerate(ordered_keys, start=1)
                ],
                columns=["번호", "결과표", "행 수"],
            )
        ),
        "",
        "## 4. 전체 분석 결과표",
        "",
    ]
    for index, key in enumerate(ordered_keys, start=1):
        title = MARKDOWN_TABLE_TITLES.get(key, key.replace("_", " "))
        frame = canonical_tables[key]
        lines.extend(
            [
                f"### 4.{index} {title}",
                "",
                f"결과 행 수: **{len(frame):,}개**",
                "",
                dataframe_to_markdown(frame),
                "",
            ]
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8-sig")


def svg_bar(
    frame: pd.DataFrame, label_column: str, value_column: str, limit: int = 10
) -> str:
    if frame.empty or label_column not in frame or value_column not in frame:
        return "<p>표시할 자료가 없습니다.</p>"
    subset = frame[[label_column, value_column]].dropna().head(limit)
    if subset.empty:
        return "<p>표시할 자료가 없습니다.</p>"
    width, row_height, left, right = 900, 34, 210, 40
    height = 35 + len(subset) * row_height
    maximum = max(float(subset[value_column].max()), 1.0)
    bars = []
    for index, (_, row) in enumerate(subset.iterrows()):
        y = 12 + index * row_height
        bar_width = (width - left - right) * float(row[value_column]) / maximum
        label = html.escape(str(row[label_column])[:28])
        value = float(row[value_column])
        bars.append(
            f'<text x="{left - 8}" y="{y + 19}" text-anchor="end">{label}</text>'
            f'<rect x="{left}" y="{y}" width="{bar_width:.1f}" height="23" rx="3" fill="#2F75B5" />'
            f'<text x="{left + bar_width + 6:.1f}" y="{y + 18}">{value:,.2f}</text>'
        )
    return (
        f'<svg role="img" viewBox="0 0 {width} {height}" width="100%" '
        f'xmlns="http://www.w3.org/2000/svg">{"".join(bars)}</svg>'
    )


def svg_line(frame: pd.DataFrame, x_column: str, value_column: str) -> str:
    subset = frame[[x_column, value_column]].dropna()
    if len(subset) < 2:
        return "<p>표시할 월별 자료가 충분하지 않습니다.</p>"
    width, height, left, right, top, bottom = 900, 300, 60, 30, 25, 55
    values = subset[value_column].astype(float).to_numpy()
    minimum, maximum = float(values.min()), float(values.max())
    span = maximum - minimum if maximum > minimum else 1.0
    points = []
    labels = []
    for index, (_, row) in enumerate(subset.iterrows()):
        x = left + index * (width - left - right) / (len(subset) - 1)
        y = top + (maximum - float(row[value_column])) * (height - top - bottom) / span
        points.append(f"{x:.1f},{y:.1f}")
        if index % max(1, len(subset) // 8) == 0 or index == len(subset) - 1:
            labels.append(
                f'<text x="{x:.1f}" y="{height - 18}" text-anchor="middle">'
                f'{html.escape(str(row[x_column]))}</text>'
            )
    return (
        f'<svg role="img" viewBox="0 0 {width} {height}" width="100%" '
        f'xmlns="http://www.w3.org/2000/svg">'
        f'<line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#999" />'
        f'<polyline points="{" ".join(points)}" fill="none" stroke="#2F75B5" stroke-width="4" />'
        + "".join(labels)
        + "</svg>"
    )


def table_html(frame: pd.DataFrame, rows: int = 15) -> str:
    if frame.empty:
        return "<p>자료가 없습니다.</p>"
    display = frame.head(rows).copy()
    for column in display.select_dtypes(include=["number"]).columns:
        display[column] = display[column].map(
            lambda value: "" if pd.isna(value) else f"{value:,.3f}"
        )
    return display.to_html(index=False, escape=True, border=0, classes="data-table")


def write_html_report(path: Path, tables: dict[str, pd.DataFrame], manifest: dict[str, Any]) -> None:
    sentiment = tables["sentiment"]
    dominant_sentiment = (
        str(sentiment.sort_values("comment_count", ascending=False).iloc[0]["sentiment"])
        if not sentiment.empty
        else "자료 없음"
    )
    topic = tables["topic"]
    top_topic = str(topic.iloc[0]["topic_primary"]) if not topic.empty else "자료 없음"
    correlation = tables["correlations"]
    view_like = correlation[
        (correlation["x"] == "view_count") & (correlation["y"] == "like_count")
    ]
    rho = (
        float(view_like.iloc[0]["spearman_rho"])
        if not view_like.empty and pd.notna(view_like.iloc[0]["spearman_rho"])
        else np.nan
    )
    generated = html.escape(manifest["generated_at_utc"])
    author_credit = html.escape(manifest.get("author_credit", AUTHOR_CREDIT))
    app_version = html.escape(manifest.get("app_version", APP_VERSION))
    demo_note = (
        '<div class="warning">현재 결과는 기능 검증용 가상 데이터입니다.</div>'
        if manifest.get("demo_data")
        else ""
    )
    document = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>유튜브 스타트업 경향 분석</title>
<style>
body {{ margin: 0; font-family: Arial, 'Malgun Gothic', sans-serif; color: #203040; background: #f4f7fa; }}
main {{ max-width: 1180px; margin: 0 auto; padding: 36px 24px 80px; }}
h1 {{ color: #17365d; margin-bottom: 8px; }} h2 {{ margin-top: 40px; color: #1f4e78; }}
.subtitle {{ color: #607080; }} .cards {{ display: grid; grid-template-columns: repeat(auto-fit,minmax(190px,1fr)); gap: 14px; margin: 24px 0; }}
.card, section {{ background: white; border-radius: 12px; padding: 20px; box-shadow: 0 3px 12px rgba(30,60,90,.08); }}
.card strong {{ display: block; font-size: 1.7rem; color: #2f75b5; margin-top: 8px; }}
.warning {{ background: #fff3cd; border-left: 5px solid #ffc107; padding: 14px; margin: 18px 0; }}
.data-table {{ border-collapse: collapse; width: 100%; font-size: .9rem; overflow: auto; display: block; }}
.data-table th {{ background: #1f4e78; color: white; }} .data-table th,.data-table td {{ padding: 8px 10px; border: 1px solid #dce4ec; white-space: nowrap; }}
svg text {{ font-size: 13px; fill: #34495e; }} .note {{ color: #6b7785; font-size: .9rem; }}
</style>
</head>
<body><main>
<h1>유튜브 ‘스타트업·startup’ 경향 분석</h1>
<p class="subtitle">{author_credit} · {app_version}</p>
<p class="subtitle">생성 시각: {generated}</p>
{demo_note}
<div class="cards">
  <div class="card">분석 영상<strong>{manifest['video_count']:,}</strong></div>
  <div class="card">분석 댓글<strong>{manifest['comment_count']:,}</strong></div>
  <div class="card">최다 영상 주제<strong>{html.escape(top_topic)}</strong></div>
  <div class="card">최다 댓글 감성<strong>{html.escape(dominant_sentiment)}</strong></div>
</div>
<section><h2>월별 영상 게시 경향</h2>{svg_line(tables['monthly_videos'], 'publish_month', 'video_count')}</section>
<section><h2>영상 주제 분포</h2>{svg_bar(tables['topic'], 'topic_primary', 'video_count')}</section>
<section><h2>댓글 빈출어</h2>{svg_bar(tables['word_comment'], 'word', 'term_frequency')}</section>
<section><h2>핵심 통계</h2>
<p>조회수와 좋아요 수의 Spearman 상관계수: <strong>{rho:.3f}</strong></p>
<p class="note">상관계수와 회귀분석은 연관성을 보여주며 인과관계를 의미하지 않습니다.</p>
{table_html(tables['descriptive'], 12)}</section>
<section><h2>검색어별 비교</h2>{table_html(tables['query'], 12)}</section>
<section><h2>댓글 감성 구성</h2>{table_html(tables['sentiment'], 10)}
<p class="note">감성은 설정 파일의 소규모 사전으로 분류한 탐색적 지표입니다.</p></section>
<section><h2>일평균 조회수 상위 영상</h2>{table_html(tables['top_videos'], 20)}</section>
<section><h2>해석상 주의</h2><ol>
<li>유튜브 검색 결과는 무작위 표본이 아닙니다.</li>
<li>조회수와 참여량은 수집 시점의 누적값입니다.</li>
<li>마지막 달이 완전하지 않으면 월별 게시량이 낮게 보일 수 있습니다.</li>
<li>댓글이 비활성화되었거나 삭제된 영상의 의견은 포함되지 않습니다.</li>
</ol></section>
</main></body></html>"""
    path.write_text(document, encoding="utf-8")


def analyze(
    raw_dir: str | Path,
    output_dir: str | Path,
    config: dict[str, Any],
    demo: bool = False,
) -> dict[str, Any]:
    raw_dir = Path(raw_dir)
    output_dir = Path(output_dir)
    tables_dir = output_dir / "tables"
    output_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)

    video_path = raw_dir / "videos.csv"
    comment_path = raw_dir / "comments.csv"
    if not video_path.exists():
        raise FileNotFoundError(f"영상 원자료가 없습니다: {video_path}")
    videos = pd.read_csv(video_path, encoding="utf-8-sig")
    comments = (
        pd.read_csv(comment_path, encoding="utf-8-sig")
        if comment_path.exists() and comment_path.stat().st_size > 0
        else pd.DataFrame()
    )
    videos = enrich_videos(videos, config)
    comments = enrich_comments(comments, config)

    iterations = max(50, safe_int(config.get("bootstrap_iterations"), 500))
    rng = np.random.default_rng(safe_int(config.get("random_seed"), 42))
    stopwords = {str(word).lower() for word in config.get("stopwords", [])}
    min_length = safe_int(config.get("min_token_length"), 2)

    monthly_videos = monthly_video_summary(videos)
    sentiment, monthly_comments = sentiment_summaries(comments)
    video_texts = (
        videos["title"].fillna("").astype(str)
        + " "
        + videos["description"].fillna("").astype(str)
    )
    top_columns = [
        "video_id",
        "title",
        "channel_title",
        "published_at",
        "matched_queries",
        "topic_primary",
        "view_count",
        "views_per_day",
        "engagement_per_1000_views",
        "video_url",
    ]
    top_videos = videos.sort_values("views_per_day", ascending=False)[top_columns].head(100)

    tables = {
        "descriptive": descriptive_summary(
            videos,
            [
                "view_count",
                "like_count",
                "comment_count",
                "duration_minutes",
                "age_days",
                "views_per_day",
                "likes_per_1000_views",
                "comments_per_1000_views",
                "engagement_per_1000_views",
            ],
        ),
        "monthly_videos": monthly_videos,
        "trend": trend_estimates(monthly_videos),
        "query": group_summary(videos, "query_group", iterations, rng),
        "topic": group_summary(videos, "topic_primary", iterations, rng),
        "duration": group_summary(videos, "duration_band", iterations, rng),
        "channels": channel_summary(videos),
        "sentiment": sentiment,
        "monthly_comments": monthly_comments,
        "correlations": spearman_with_ci(
            videos,
            [
                ("view_count", "like_count"),
                ("view_count", "comment_count"),
                ("views_per_day", "engagement_per_1000_views"),
                ("channel_subscriber_count", "views_per_day"),
                ("duration_minutes", "engagement_per_1000_views"),
            ],
            iterations,
            rng,
        ),
        "regression": bootstrap_regression(videos, iterations, rng),
        "word_video": word_frequency(video_texts, stopwords, min_length),
        "word_comment": word_frequency(
            comments["text"] if not comments.empty else [], stopwords, min_length
        ),
        "cooccurrence": cooccurrence_table(
            comments["text"] if not comments.empty else [], stopwords, min_length
        ),
        "top_videos": top_videos,
        "quality": data_quality_table(videos, comments),
        "videos": videos,
        "comments": comments,
    }

    advanced_keys: list[str] = []
    if run_advanced_statistics is not None:
        advanced_tables = run_advanced_statistics(videos, comments, config)
        tables.update(advanced_tables)
        advanced_keys = list(advanced_tables)
    else:
        tables["method_catalog"] = pd.DataFrame(
            [
                {
                    "category": "고급 통계",
                    "method": "실행 불가",
                    "purpose": "scipy, statsmodels, scikit-learn 설치 필요",
                    "status": "필수 라이브러리 미설치",
                }
            ]
        )
        advanced_keys = ["method_catalog"]

    table_files = {
        "descriptive": "descriptive_statistics.csv",
        "monthly_videos": "monthly_video_trends.csv",
        "trend": "trend_estimates.csv",
        "query": "query_comparison.csv",
        "topic": "topic_summary.csv",
        "duration": "duration_summary.csv",
        "channels": "channel_summary.csv",
        "sentiment": "sentiment_summary.csv",
        "monthly_comments": "monthly_comment_trends.csv",
        "correlations": "spearman_correlations.csv",
        "regression": "view_regression.csv",
        "word_video": "video_word_frequency.csv",
        "word_comment": "comment_word_frequency.csv",
        "cooccurrence": "comment_cooccurrence.csv",
        "top_videos": "top_videos.csv",
        "quality": "data_quality.csv",
    }
    for key, filename in table_files.items():
        save_csv(tables[key], tables_dir / filename)
    for key in advanced_keys:
        save_csv(tables[key], tables_dir / f"advanced_{key}.csv")
    save_csv(videos.drop(columns=["published_datetime"], errors="ignore"), output_dir / "videos_enriched.csv")
    save_csv(comments.drop(columns=["published_datetime"], errors="ignore"), output_dir / "comments_enriched.csv")

    manifest = {
        "generated_at_utc": utc_now_iso(),
        "author_credit": AUTHOR_CREDIT,
        "app_version": APP_VERSION,
        "demo_data": bool(demo),
        "video_count": len(videos),
        "comment_count": len(comments),
        "unique_channel_count": int(videos["channel_id"].nunique()),
        "queries": config["queries"],
        "published_after": config.get("published_after"),
        "published_before": config.get("published_before"),
        "bootstrap_iterations": iterations,
        "advanced_methods": config.get("advanced_methods", "all_applicable"),
        "advanced_table_count": len(advanced_keys),
        "raw_directory": str(raw_dir.resolve()),
    }
    (output_dir / "run_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_excel_report(output_dir / "startup_youtube_analysis.xlsx", tables, config)
    write_html_report(output_dir / "startup_youtube_report.html", tables, manifest)
    write_markdown_report(
        output_dir / "startup_youtube_all_results.md", tables, manifest, config
    )
    return manifest


def create_demo_data(raw_dir: str | Path, config: dict[str, Any]) -> None:
    raw_dir = Path(raw_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)
    rng = np.random.default_rng(safe_int(config.get("random_seed"), 42))
    random.seed(safe_int(config.get("random_seed"), 42))
    collected_at = datetime.now(timezone.utc)
    topics = {
        "투자·자금": ["스타트업 투자유치 전략", "Startup seed funding story", "VC가 보는 창업팀"],
        "창업·경영": ["스타트업 창업자가 말하는 사업계획", "Startup founder interview", "초기 창업 경영 노하우"],
        "기술·제품": ["AI 스타트업 제품 만들기", "SaaS startup product demo", "기술 스타트업 플랫폼 분석"],
        "성장·마케팅": ["스타트업 성장과 고객 마케팅", "Startup growth strategy", "매출을 만드는 스케일업"],
        "실패·위기": ["스타트업 실패에서 배운 것", "Why startups fail", "폐업과 위기를 넘는 방법"],
        "정책·생태계": ["정부 창업지원 정책", "Korean startup ecosystem", "액셀러레이터와 창업 생태계"],
    }
    channels = [
        ("UC_DEMO_A", "창업연구소", 120000),
        ("UC_DEMO_B", "비즈니스인사이트", 68000),
        ("UC_DEMO_C", "Startup Korea", 35000),
        ("UC_DEMO_D", "테크리뷰", 210000),
        ("UC_DEMO_E", "대표의하루", 18000),
        ("UC_DEMO_F", "Venture Talk", 92000),
    ]
    topic_effect = {
        "투자·자금": 1.35,
        "창업·경영": 1.0,
        "기술·제품": 1.45,
        "성장·마케팅": 1.2,
        "실패·위기": 1.3,
        "정책·생태계": 0.75,
    }
    video_records: list[dict[str, Any]] = []
    comment_records: list[dict[str, Any]] = []
    months_ago = 18
    for index in range(72):
        topic = list(topics)[index % len(topics)]
        title = topics[topic][index % len(topics[topic])]
        channel_id, channel_title, subscribers = channels[index % len(channels)]
        days_ago = int(rng.integers(5, months_ago * 30))
        published = collected_at - timedelta(days=days_ago, hours=int(rng.integers(0, 24)))
        duration = int(rng.choice([45, 180, 420, 780, 1500, 2700]))
        baseline_daily = max(20, subscribers / 60) * topic_effect[topic]
        daily_views = baseline_daily * float(rng.lognormal(0, 0.75))
        views = int(max(50, daily_views * days_ago))
        engagement_rate = float(np.clip(rng.normal(28, 9), 4, 70))
        likes = int(views * engagement_rate / 1000 * 0.82)
        comment_count = max(2, int(views * engagement_rate / 1000 * 0.18))
        if "Startup" in title or "startup" in title or "startups" in title:
            matched = "startup"
        elif index % 7 == 0:
            matched = "스타트업|startup"
        else:
            matched = "스타트업"
        video_id = f"DEMO{index:06d}"
        video_records.append(
            {
                "video_id": video_id,
                "title": title,
                "description": f"{title}. {topic} 관련 사례와 데이터, 현장 경험을 설명합니다.",
                "published_at": published.isoformat().replace("+00:00", "Z"),
                "channel_id": channel_id,
                "channel_title": channel_title,
                "tags": f"스타트업|startup|{topic}",
                "category_id": "27",
                "default_language": "ko",
                "default_audio_language": "ko",
                "duration_iso": "",
                "duration_seconds": duration,
                "caption_available": "true",
                "definition": "hd",
                "view_count": views,
                "like_count": likes,
                "comment_count": comment_count,
                "privacy_status": "public",
                "made_for_kids": False,
                "video_url": f"https://www.youtube.com/watch?v={video_id}",
                "matched_queries": matched,
                "search_ranks_json": "{}",
                "channel_subscriber_count": subscribers,
                "channel_hidden_subscriber_count": False,
                "channel_view_count": subscribers * 250,
                "channel_video_count": int(rng.integers(60, 800)),
                "collected_at_utc": collected_at.isoformat().replace("+00:00", "Z"),
            }
        )

        templates = [
            ("좋은 내용이고 정말 유익합니다. 스타트업 창업에 도움이 됐어요", "긍정"),
            ("혁신적인 사례네요. 성공과 성장이 기대됩니다", "긍정"),
            ("투자와 매출 자료를 조금 더 보고 싶습니다", "중립"),
            ("설명 감사합니다. 창업자 인터뷰가 궁금합니다", "긍정"),
            ("스타트업 거품과 과장이 너무 심한 것 같습니다", "부정"),
            ("실패 사례와 위기 대응도 다뤄주세요", "부정"),
            ("정부지원 정책의 실제 효과가 궁금합니다", "중립"),
            ("Great startup story and helpful advice", "긍정"),
            ("This looks like hype and the business model has a problem", "부정"),
            ("SaaS product and customer growth data please", "중립"),
        ]
        demo_comment_count = int(rng.integers(4, 11))
        for comment_index in range(demo_comment_count):
            text, _ = templates[(index + comment_index) % len(templates)]
            comment_id = f"C{index:04d}_{comment_index:03d}"
            is_reply = comment_index > 0 and comment_index % 5 == 0
            comment_records.append(
                {
                    "comment_id": comment_id,
                    "video_id": video_id,
                    "parent_id": f"C{index:04d}_000" if is_reply else "",
                    "is_reply": is_reply,
                    "source": "inline_reply" if is_reply else "top_level",
                    "text": text,
                    "like_count": int(rng.integers(0, 45)),
                    "published_at": (
                        published + timedelta(days=int(rng.integers(0, max(days_ago, 1))))
                    ).isoformat().replace("+00:00", "Z"),
                    "updated_at": "",
                    "collected_at_utc": collected_at.isoformat().replace("+00:00", "Z"),
                }
            )

    save_csv(pd.DataFrame(video_records), raw_dir / "videos.csv")
    save_csv(pd.DataFrame(comment_records), raw_dir / "comments.csv")
    save_csv(
        pd.DataFrame(columns=["video_id", "stage", "status", "reason", "message"]),
        raw_dir / "collection_errors.csv",
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="유튜브 스타트업·startup 영상 및 댓글 경향 분석"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    for name in ["collect", "all", "demo"]:
        subparser = subparsers.add_parser(name)
        subparser.add_argument("--config", default="config.json", help="설정 JSON 파일")
        subparser.add_argument("--out", required=True, help="실행 결과 폴더")
        if name in {"collect", "all"}:
            subparser.add_argument(
                "--api-key-env",
                default="YOUTUBE_API_KEY",
                help="API 키가 저장된 환경변수 이름",
            )

    analyze_parser = subparsers.add_parser("analyze")
    analyze_parser.add_argument("--config", default="config.json", help="설정 JSON 파일")
    analyze_parser.add_argument("--input", required=True, help="videos.csv와 comments.csv 폴더")
    analyze_parser.add_argument("--out", required=True, help="분석 결과 폴더")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config = load_config(args.config)
    try:
        if args.command == "collect":
            result = collect(args.out, config, os.environ.get(args.api_key_env, ""))
        elif args.command == "analyze":
            result = analyze(args.input, args.out, config)
        elif args.command == "all":
            collect(args.out, config, os.environ.get(args.api_key_env, ""))
            result = analyze(Path(args.out) / "raw", Path(args.out) / "analysis", config)
        elif args.command == "demo":
            root = Path(args.out)
            create_demo_data(root / "raw", config)
            result = analyze(root / "raw", root / "analysis", config, demo=True)
        else:
            raise AssertionError("알 수 없는 명령")
    except (ValueError, FileNotFoundError, YouTubeAPIError) as exc:
        print(f"오류: {exc}", file=sys.stderr)
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
